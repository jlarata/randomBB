import shelve
import openpyxl
from random import randrange
import os
import time

# openpyxl configuration

# path of the folder where the xlsx is
path = "C:/.../bibioteca.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row


# 'libro' (book) object and list initialization.

class Libro:
    def __init__(self, numero, autor, titulo, genero, seccion, isLeido):
        self.numero = numero
        self.autor = autor
        self.titulo = titulo
        self.genero = genero
        self.seccion = seccion
        self.isLeido = isLeido

# bilioteca is used the first time to create de database. Read lines from 239
biblioteca = []
BBProvisoriaDeIsLeidos = []
yes_choices = ['yes', 'y', 'YES', 'Y', 'Yes', 'Si', 'S', 's', 'si']

def print_slow(str):
    for letter in str:
        print(letter, end='', flush=True)
        time.sleep(0.005)

    
######################################################    
################## Shelve and Excel ##################
######################################################

def abreLaShelve():
    # writeback=true: allegedly unnecesary and slowing-down the program. 
    # neverdeless, if it was not put there, i couldn't modify the shelve
    my_shelve = shelve.open("mydata.db", writeback=True)
    return my_shelve

def pruebaShelveContraExcel():
    
    # these are just a bunch of useful methods for debugging 
    # print("len of myshelve[bibioteca]: ", len(my_shelve["biblioteca"]))
    # print("last element of myshelve: ", my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].numero, " | ", my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].autor, my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].titulo, " | ", my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].genero, " | ", my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].seccion, " | ", my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].isLeido)
    # print("rows of excel: ",  m_row)
    # print("last record of excel: ",  sheet_obj.cell(row = m_row, column = 3).value)
    
    # i'll leave this method here as example of how to edit a register in the shelve
    # my_shelve["biblioteca"][len(my_shelve["biblioteca"])-1].seccion = "Literatura estadounidense" 
    
    
    # boolean: checks if records are uneven bewteen excel and shelve
    hayNuevosLibros = (len(my_shelve["biblioteca"]) != m_row)
    
    # if not, that means used added new records to excel
    if hayNuevosLibros:

        # (int) establish the amount of new records
        cantidadNueva = m_row - len(my_shelve["biblioteca"])
        # alert the user and print detailed list:
        print_slow(f"Atención!! se ha detectado que el excel tiene {cantidadNueva} nuevo libro: \n\n") if cantidadNueva == 1 else print_slow(f"Atención!! se ha detectado que el excel tiene {cantidadNueva} nuevos libros: \n\n")
        aux = 1
        for x in range(cantidadNueva):
            mensaje = str(sheet_obj.cell(row = (m_row-cantidadNueva+aux), column = 2).value + sheet_obj.cell(row = (m_row-cantidadNueva+aux), column = 3).value + "\n")
            print_slow(mensaje)
            aux += 1
    
        # gives user the chance of update database or go ahead without adding new records
        ingresar = input("\n Si querés que los agreguemos a la base de datos ahora, ingresa 'Y': \n")
        if ingresar in yes_choices:
            os.system('cls')
            print_slow("\n Ok... \n\n")
            agregaLibros(cantidadNueva)
        else:
            os.system('cls')
            print_slow("\n Ok, continuemos entonces... \n")

def agregaLibros(cantidadNueva):
    # appends as many new objects as "cantidad nueva".
    # the iteration method/arguments are to obtain the "self.numero" column and start there
    for i in range(m_row-cantidadNueva+1, m_row+1):
        my_shelve["biblioteca"].append(Libro(i, "autor", "titulo", "genero", "seccion", False))
        
    # then using again the same itration method/arguments... (probably could fusion both iterations)
    # 1. declare values from excel cells as variables
    for i in range(m_row-cantidadNueva+1, m_row+1):
        autor = sheet_obj.cell(row = i, column = 2)
        titulo = sheet_obj.cell(row = i, column = 3)
        genero = sheet_obj.cell(row = i, column = 4)
        seccion = sheet_obj.cell(row = i, column = 5)
        
        # 2. use this variables as attributes of the new objects
        
        # index is -1 because "i" here refers to the self.numero column, which is not in base 0 but 1.
        # case the user didn't fill one of the columns, the program fills that column with "..." 
        if(autor.value):
            my_shelve["biblioteca"][i-1].autor = str(autor.value)
        else:
            my_shelve["biblioteca"][i-1].autor = "..."
        if(titulo.value):
            my_shelve["biblioteca"][i-1].titulo = str(titulo.value)
        else:
            my_shelve["biblioteca"][i-1].titulo = "..."
        if(genero.value):
            my_shelve["biblioteca"][i-1].genero = str(genero.value)
        else:
            my_shelve["biblioteca"][i-1].genero = "..."
        if(seccion.value):
            my_shelve["biblioteca"][i-1].seccion = str(seccion.value)
        else:
            my_shelve["biblioteca"][i-1].seccion = "..."
    
    # print when success
    print_slow(f"Se ha agregado con éxito el siguiente libro: \n\n") if cantidadNueva == 1 else print_slow(f"Se han agregado con éxito los siguientes libros: \n\n")
    aux = 1
    for i in range(m_row-cantidadNueva+1, m_row+1):
        mensaje = str(str(my_shelve["biblioteca"][i-1].numero) + " | " + my_shelve["biblioteca"][i-1].autor + my_shelve["biblioteca"][i-1].titulo + " | " + my_shelve["biblioteca"][i-1].genero + " | " + my_shelve["biblioteca"][i-1].seccion + " \n")
        print_slow(mensaje)
        aux += 1
    
    print_slow("\n\n Presione cualquier tecla para continuar... \n\n")
    input("")

######################################################    
################# General Functions ##################
######################################################


def saludaAlUsuario():
    print_slow("\n Hola, vamos a usar un número al azar para elegir un libro de la biblioteca, presiona enter para continuar \n")
    input("")

# returns random Number
def eligeNumeroAlAzar():
    numeroAlAzar = randrange(len(my_shelve["biblioteca"])+1)
    return numeroAlAzar
    
# returns random book object (Libro)
def eligeLibroAlAzar(numeroAlAzar):
    libroAlAzar = my_shelve["biblioteca"][numeroAlAzar]
    return libroAlAzar


# recursive: if the book is read (.isLeido), append it to a temporary library (BBProvisoria) and choose
# another random number and book.
# otherwise, continue with the program
def separaIsLeidos(numeroAlAzar, libroAlAzar):
    if libroAlAzar.isLeido:
        enviaTodosLosIsLeidoABBProvisoriaDeIsLeidos(libroAlAzar)
    else:
        comunicaLibroAsignadoAlAzar(numeroAlAzar, libroAlAzar)

def enviaTodosLosIsLeidoABBProvisoriaDeIsLeidos(libroAlAzar):
    BBProvisoriaDeIsLeidos.append(libroAlAzar) 
    numeroAlAzar = eligeNumeroAlAzar()
    libroAlAzar = eligeLibroAlAzar(numeroAlAzar)
    separaIsLeidos(numeroAlAzar, libroAlAzar)

# prints:
def comunicaLibroAsignadoAlAzar(numeroAlAzar, libroAlAzar):
    #1: already read ("isLeido") books (BBProvisoriaDeIsLeidos)
    comunicaLosLibrosLeidosQueSalieron()
    #2: current random number and random book w/ detail:
    print_slow("\n ...salió sorteado el número "+str(numeroAlAzar)+", al que le corresponde el siguiente libro: \n")
    print_slow(f"""{libroAlAzar.autor}, {libroAlAzar.titulo}, {libroAlAzar.genero}, {libroAlAzar.seccion}\n""")

# CASE A: first random book has not been read (isLeido = False)
# Recursive: The user accepts to read this book or choose another.
def aceptaLibroOPideOtro(libroAlAzar):
    indicaSiVaALeer = input("\n Según la base de datos, aún no has leído este libro. ¿Vas a leerlo ahora? Y/N: ")
    if indicaSiVaALeer in yes_choices:
        print_slow("\n ¡Excelente! Que lo disfrutes.\n")
        modificaLibroIsLeido(libroAlAzar)
    else:
        os.system('cls')
        print("\n bueno, elijamos otro...\n")
        BBProvisoriaDeIsLeidos.clear()
        numeroAlAzar = eligeNumeroAlAzar()
        libroAlAzar = eligeLibroAlAzar(numeroAlAzar)
        separaIsLeidos(numeroAlAzar, libroAlAzar)
        aceptaLibroOPideOtro(libroAlAzar)

# CASE B: At least one book already read was first selected.
def comunicaLosLibrosLeidosQueSalieron():
    if any(BBProvisoriaDeIsLeidos):
        print_slow("\n ...bueno, primero salieron estos libros, que ya leiste \n")
        for libro in BBProvisoriaDeIsLeidos:
            print("   -> "+libro.titulo)
        detieneLaFuncionSiHayErrorEnIsLeidos()   

# this function stops the program if user wants to read a book in the temporary list of already read book
# (maybe a re-read or maybe a wrong record). no need for futher alteration because the output alreay is 
# the expected: random book has isLeido atribute as True.
def detieneLaFuncionSiHayErrorEnIsLeidos():
    errorEnIsLeidos = input("\n Si entre estos libros hay uno que querías leer ahora, ingresa 'Y' para detener el programa: \n")
    if errorEnIsLeidos in yes_choices:
        print_slow("\n Ok, podés leer ese libro entonces, que lo disfrutes. \n")
        cierraLaShelve()
        despideAlUsuario()
        cierraElPrograma()

# Update column "isLeido" (already read) to True
def modificaLibroIsLeido(libroAlAzar):
    my_shelve["biblioteca"][libroAlAzar.numero-1].isLeido = True
    print(f'\n Modificada la condición de "{my_shelve["biblioteca"][libroAlAzar.numero-1].titulo}" a leido: {my_shelve["biblioteca"][libroAlAzar.numero-1].isLeido}')

def cierraLaShelve():
    my_shelve.close()
    
def despideAlUsuario():
    input("\n Cerrando el programa, espero que te haya servido. presiona enter para terminar.\n")
    cierraElPrograma()

def cierraElPrograma():
    exit(0)

######################################################    
######## Creating and populating the Database ########
############## using data from a xlsx ################
######################################################

def crearBiblioteca():
    if (biblioteca): {
        print('hay biblioteca')
    }
    else: {
        print('no hay biblioteca!')
    }
        
    # assigns to shelve
    my_shelve["biblioteca"] = biblioteca

    for i in range(m_row):

        biblioteca.append(Libro(i+1, "autor", "titulo", "genero", "seccion", False))
                
        autor = sheet_obj.cell(row = i+1, column = 2)
        titulo = sheet_obj.cell(row = i+1, column = 3)
        genero = sheet_obj.cell(row = i+1, column = 4)
        seccion = sheet_obj.cell(row = i+1, column = 5)
        
        if(autor.value):
            my_shelve["biblioteca"][i-1].autor = str(autor.value)
        else:
            my_shelve["biblioteca"][i-1].autor = "..."
        if(titulo.value):
            my_shelve["biblioteca"][i-1].titulo = str(titulo.value)
        else:
            my_shelve["biblioteca"][i-1].titulo = "..."
        if(genero.value):
            my_shelve["biblioteca"][i-1].genero = str(genero.value)
        else:
            my_shelve["biblioteca"][i-1].genero = "..."
        if(seccion.value):
            my_shelve["biblioteca"][i-1].seccion = str(seccion.value)
        else:
            my_shelve["biblioteca"][i-1].seccion = "..."

    

    # prints it all
    for libro in my_shelve["biblioteca"]:
        print(str(libro.numero)+", "+(libro.autor)+", "+(libro.titulo)+", "+(libro.genero)+", "+(libro.seccion)+", "+str(libro.isLeido))

if __name__ == "__main__":
    my_shelve = abreLaShelve()
    crearBiblioteca()
    saludaAlUsuario()
    pruebaShelveContraExcel()
    numeroAlAzar = eligeNumeroAlAzar()
    libroAlAzar = eligeLibroAlAzar(numeroAlAzar)
    separaIsLeidos(numeroAlAzar, libroAlAzar)
    aceptaLibroOPideOtro(libroAlAzar)
    cierraLaShelve()
    despideAlUsuario()
    cierraElPrograma()

    



